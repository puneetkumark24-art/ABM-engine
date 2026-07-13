"""
abm_engine/core/loader.py
──────────────────────────
Reads your ABM Excel and loads contacts + accounts into the database.
Handles all new fields: persona, segment, is_ksa_national, has_warm_relationship.
"""
from __future__ import annotations
from pathlib import Path
from loguru import logger
import openpyxl

from .models import Contact, Tier, Persona, Segment, AccountType
from ..database.db import upsert_contact, upsert_account

# KSA national name indicators — used to auto-detect is_ksa_national
ARABIC_NAME_PATTERNS = [
    "Al ", "Al-", "Bin ", "Ibn ", "Abdullah", "Mohammed", "Ahmad",
    "Khalid", "Sultan", "Faisal", "Saud", "Turki", "Mansour",
    "Nasser", "Fahad", "Waleed", "Saleh", "Omar", "Ali",
]

# Persona detection from role string
PERSONA_KEYWORDS = {
    # Order matters — first match wins. More specific patterns first.
    Persona.CISO:            ["ciso", "chief information security", "head of security", "cybersecurity", "information security"],
    Persona.CEO:             ["ceo", "chief executive officer", "managing director", "general manager", "co-founder", "founder", "president"],
    Persona.CDO:             ["cdo", "chief digital officer", "head of digital", "vp digital", "digital officer"],
    Persona.CTO:             ["cto", "chief technology officer", "chief technical", "head of technology", "vp technology", "head of it"],
    Persona.HEAD_RETAIL:     ["retail banking", "consumer banking", "head of retail", "head of consumer", "personal banking"],
    Persona.HEAD_PRODUCT:    ["head of product", "vp product", "chief product officer", "product officer"],
    Persona.HEAD_COMPLIANCE: ["compliance", "head of risk", "regulatory", "chief risk", "audit"],
}

SEGMENT_KEYWORDS = {
    Segment.DIGITAL:    ["digital bank", "neobank", "stc bank", "d360", "vision bank"],
    Segment.BNPL:       ["bnpl", "tamara", "tabby", "buy now"],
    Segment.SME:        ["sme", "lendo", "funding souq", "erad", "small medium"],
    Segment.EMBEDDED:   ["embedded", "hala", "platform"],
    Segment.PAYMENTS:   ["payment", "geidea", "hyperpay", "emtech"],
    Segment.ISLAMIC:    ["islamic", "al rajhi", "albilad", "alinma", "aljazira", "aljazira"],
    Segment.COMMERCIAL: ["commercial", "snb", "riyad", "sabb", "fransi", "arab national"],
}

FI_INSTITUTIONS = [
    "tamara", "tabby", "hala", "lendo", "erad", "lean technologies",
    "funding souq", "geidea", "hyperpay", "manafa", "scopeer",
    "abdul latif jameel", "saudi real estate refinance",
]

COLUMN_MAP = {
    "full_name":        ["Full Name", "Name"],
    "role":             ["Current Role", "Role", "Title", "Job Title"],
    "seniority":        ["Seniority", "Level"],
    "institution":      ["Institution", "Company", "Bank"],
    "country":          ["Country"],
    "institution_type": ["Type", "Institution Type"],
    "email":            ["Work Email (Pattern)", "Email", "Work Email"],
    "email_confidence": ["Email Confidence", "Confidence"],
    "linkedin_url":     ["LinkedIn URL", "LinkedIn"],
    "hq_phone":         ["HQ Phone", "Phone"],
    "key_signal":       ["Key Business Signal (Verified)", "Signal", "Key Signal", "Key Business Signal"],
    "outreach_angle":   ["Outreach Angle", "Angle"],
    "product_fit":      ["Decimal Product Fit", "Product Fit"],
    "warmness":         ["Warmness", "Warmness Indicator"],
    "priority_score":   ["Priority Score", "Score", "Priority Score (1–100)"],
    "is_ksa_national":  ["KSA National", "Is KSA National", "National"],
    "has_warm_relationship": ["Warm Relationship", "Existing Contact", "Warm"],
    "segment":          ["Segment"],
    "persona":          ["Persona"],
}


def _find_col(headers: list, options: list) -> int | None:
    for opt in options:
        for i, h in enumerate(headers):
            if h and opt.lower() in str(h).lower():
                return i
    return None


def _detect_persona(role: str) -> str:
    role_lower = role.lower()
    for persona, keywords in PERSONA_KEYWORDS.items():
        if any(k in role_lower for k in keywords):
            return persona.value
    return Persona.OTHER.value


def _detect_segment(institution: str, institution_type: str) -> str:
    inst_lower = (institution + " " + institution_type).lower()
    for segment, keywords in SEGMENT_KEYWORDS.items():
        if any(k in inst_lower for k in keywords):
            return segment.value
    return Segment.COMMERCIAL.value


def _detect_account_type(institution: str) -> str:
    if any(fi in institution.lower() for fi in FI_INSTITUTIONS):
        return AccountType.FI.value
    return AccountType.BANK.value


def _detect_ksa_national(name: str) -> bool:
    return any(p.lower() in name.lower() for p in ARABIC_NAME_PATTERNS)


def _score_to_tier(score: int) -> str:
    if score >= 75:
        return Tier.HOT.value
    if score >= 50:
        return Tier.WARM.value
    return Tier.COLD.value


def load_contacts_from_excel(path: str | Path, sheet_name: str = None) -> int:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Auto-detect sheet
    if sheet_name:
        ws = wb[sheet_name]
    else:
        target = None
        for name in wb.sheetnames:
            if any(k in name.lower() for k in ["contact", "database", "lead"]):
                target = name
                break
        ws = wb[target or wb.sheetnames[0]]

    logger.info("Loading from sheet: '{}'", ws.title)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    # Find header row
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        if sum(1 for v in row if v is not None) > 4:
            header_row = [str(v).strip() if v else "" for v in row]
            header_idx = i
            break

    if not header_row:
        raise ValueError("Could not find header row in Excel")

    col_map = {field: _find_col(header_row, options)
               for field, options in COLUMN_MAP.items()}

    loaded = skipped = 0

    for row in rows[header_idx + 1:]:
        if not row or all(v is None for v in row):
            continue

        def get(field: str) -> str:
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        name = get("full_name")
        inst = get("institution")
        if not name or not inst:
            skipped += 1
            continue

        try:
            score = int(float(get("priority_score"))) if get("priority_score") else 50
        except ValueError:
            score = 50

        role    = get("role") or "Decision Maker"
        segment = get("segment") or _detect_segment(inst, get("institution_type") or "")
        persona = get("persona") or _detect_persona(role)

        # is_ksa_national: from column or auto-detect from name
        is_ksa_raw = get("is_ksa_national").lower()
        is_ksa     = (is_ksa_raw in ("yes", "1", "true")) or _detect_ksa_national(name)

        # warm relationship: from column
        warm_raw = get("has_warm_relationship").lower()
        warm     = warm_raw in ("yes", "1", "true")

        # Upsert account first
        acct_type = _detect_account_type(inst)
        account_id = upsert_account(
            name         = inst,
            account_type = acct_type,
            segment      = segment,
            country      = get("country") or "Saudi Arabia",
        )

        contact = Contact(
            account_id       = account_id,
            full_name        = name,
            role             = role,
            persona          = persona,
            seniority        = get("seniority") or "VP",
            is_ksa_national  = is_ksa,
            institution      = inst,
            country          = get("country") or "Saudi Arabia",
            institution_type = get("institution_type") or acct_type,
            segment          = segment,
            email            = get("email") or None,
            email_confidence = get("email_confidence") or "Low",
            linkedin_url     = get("linkedin_url") or None,
            hq_phone         = get("hq_phone") or None,
            key_signal       = get("key_signal") or "GCC digital banking expansion",
            outreach_angle   = get("outreach_angle") or "Digital banking infrastructure",
            product_fit      = get("product_fit") or "Account opening, digital lending",
            warmness         = get("warmness") or "Cold",
            has_warm_relationship = warm,
            priority_score   = score,
            tier             = _score_to_tier(score),
        )

        upsert_contact(contact)
        loaded += 1

    logger.info("Loaded {} contacts, {} skipped", loaded, skipped)
    return loaded
