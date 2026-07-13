"""
import_incoming.py — loads whatever's sitting in incoming/contacts/ and
incoming/offerings/ into the DRIP database.

Usage:
    python etl/import_incoming.py

Handles two contact-file shapes, auto-detected per sheet:
  1. Flat tables: header row is row 1, one row per person.
  2. "LinkedIn Priority Contacts" exports (the real shape Puneet's files use):
     row 1 = title, row 2 = source description, row 3 = real header, then
     data rows interrupted by "TIER N - ..." section-divider rows. LinkedIn
     URLs are stored as a hyperlink on a "View Profile"/Name cell, not as
     plain text — extracted via the cell's hyperlink target.

Institution is taken from a "Company"-style column when the sheet has one
with real values; otherwise falls back to a per-filename mapping (needed
because some exports use their "Location" column for city, not company).

Safe to re-run — matches on (full_name, institution) for people and on name
for products, so re-running after adding more rows updates existing records
instead of duplicating them.
"""
from __future__ import annotations
import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from database import Base, engine, SessionLocal
import models

HERE = Path(__file__).parent.parent
CONTACTS_DIR = HERE / "incoming" / "contacts"
OFFERINGS_DIR = HERE / "incoming" / "offerings"

# Known bank-export filenames -> canonical organization name (matches what's
# already loaded from decimal_abm where applicable). Matched case-insensitively
# against the filename stem. Extend this as you add more bank files.
FILENAME_INSTITUTION_MAP = {
    "al rajhi all contacts linkdin": "Al Rajhi Bank",
    "alinma_linkedin_priority_contacts": "Alinma Bank",
    "anb_linkedin_priority_contacts": "Arab National Bank",
    "bsf master": "Banque Saudi Fransi",
    "ejada_linkedin_priority_contacts": "EJADA",
    "riyad": "Riyad Bank",
    "snb_linkedin_priority_contacts": "Saudi National Bank (SNB)",
}

NAME_ALIASES = {"name", "full_name", "contact_name", "person"}
COMPANY_ALIASES = {"company", "institution", "organization", "organisation", "org", "bank",
                    "company_(as_shown)", "company(as_shown)"}
TITLE_ALIASES = {"title", "role", "designation", "title_/_headline", "title/headline",
                  "job_title", "current_title", "headline"}
TIER_ALIASES = {"tier"}
NOTES_ALIASES = {"note", "notes", "reason_/_priority_rationale", "reason", "background_notes"}
LINKEDIN_COL_ALIASES = {"linkedin_profile", "linkedin_url", "linkedin", "profile", "view_profile"}
LOCATION_ALIASES = {"location"}
EMAIL_ALIASES = {"email", "primary_email", "work_email"}
PHONE_ALIASES = {"phone", "mobile", "phone_number", "phone_no", "phone_no.", "contact_number"}
WHATSAPP_ALIASES = {"whatsapp", "whatsapp_number"}
SENIORITY_ALIASES = {"seniority", "seniority_level"}
PERSONA_ALIASES = {"persona", "role_type"}
COUNTRY_ALIASES = {"country"}

TIER_PRIORITY_SCORE = {"tier 1": 90, "tier 2": 60, "tier 3": 30}

# Title-text -> seniority/role heuristics. These files have no seniority
# column, only a free-text "Title / Headline" — infer from common patterns
# so the dashboard's C-Suite/Champions sections have real data instead of
# being empty. Best-effort, not authoritative; safe to hand-correct later
# via the API (PATCH not built yet — direct DB edit or a future endpoint).
C_SUITE_PATTERNS = ["chief ", "ceo", "cfo", "cto", "coo", "cro", "cio", "ciso", "chairman", "president"]
SVP_EVP_PATTERNS = ["evp", "svp", "executive vice president", "senior vice president", "group head"]
DIRECTOR_PATTERNS = ["head of", "director", "vp ", "vice president", "division head"]


def infer_seniority(title: str | None) -> str | None:
    if not title:
        return None
    t = title.lower()
    if any(p in t for p in C_SUITE_PATTERNS):
        return "c_suite"
    if any(p in t for p in SVP_EVP_PATTERNS):
        return "svp_evp"
    if any(p in t for p in DIRECTOR_PATTERNS):
        return "director"
    return None


def infer_flags(title: str | None, seniority: str | None) -> dict:
    is_dm = seniority in ("c_suite", "svp_evp")
    is_infl = seniority in ("c_suite", "svp_evp", "director")
    return {"is_decision_maker": is_dm, "is_influencer": is_infl}



def _slug(cell) -> str:
    return str(cell).strip().lower().replace(" ", "_") if cell is not None else ""


def _institution_for_file(path: Path) -> str | None:
    stem = path.stem.strip().lower()
    for key, org_name in FILENAME_INSTITUTION_MAP.items():
        if key in stem:
            return org_name
    return None


def _find_header_row(rows: list[tuple], max_scan: int = 10) -> int | None:
    for i, row in enumerate(rows[:max_scan]):
        slugs = {_slug(c) for c in row if c is not None}
        if slugs & NAME_ALIASES:
            return i
    return None


def _col_map(header_row: tuple) -> dict[str, int]:
    m = {}
    for idx, cell in enumerate(header_row):
        s = _slug(cell)
        if not s:
            continue
        if s in NAME_ALIASES and "name" not in m:
            m["name"] = idx
        elif s in COMPANY_ALIASES and "company" not in m:
            m["company"] = idx
        elif s in TITLE_ALIASES and "title" not in m:
            m["title"] = idx
        elif s in TIER_ALIASES and "tier" not in m:
            m["tier"] = idx
        elif s in NOTES_ALIASES and "notes" not in m:
            m["notes"] = idx
        elif s in LINKEDIN_COL_ALIASES and "linkedin" not in m:
            m["linkedin"] = idx
        elif s in LOCATION_ALIASES and "location" not in m:
            m["location"] = idx
        elif s in EMAIL_ALIASES and "email" not in m:
            m["email"] = idx
        elif s in PHONE_ALIASES and "phone" not in m:
            m["phone"] = idx
        elif s in WHATSAPP_ALIASES and "whatsapp" not in m:
            m["whatsapp"] = idx
        elif s in SENIORITY_ALIASES and "seniority" not in m:
            m["seniority"] = idx
        elif s in PERSONA_ALIASES and "persona" not in m:
            m["persona"] = idx
        elif s in COUNTRY_ALIASES and "country" not in m:
            m["country"] = idx
    return m


def _is_divider_row(row: tuple) -> bool:
    """A 'TIER 1 - Priority (...)' section-divider row: only the first cell has content."""
    non_empty = [c for c in row if c not in (None, "")]
    return len(non_empty) == 1 and row[0] not in (None, "")


def _iter_sheet_contacts(ws, filename_institution: str | None, source_label: str):
    # "Indians at X" / "X - India Located" sheets are how these exports flag
    # Indian-origin contacts — "india" is a substring of both "Indian" and
    # "India", so one check catches both sheet-naming patterns.
    is_indian_sheet = "india" in (ws.title or "").lower()
    rows = list(ws.iter_rows(values_only=False))
    plain_rows = [[c.value for c in r] for r in rows]
    header_idx = _find_header_row(plain_rows)
    if header_idx is None:
        return  # not a contacts sheet (e.g. a "Legend" / methodology sheet) — skip
    cmap = _col_map(tuple(plain_rows[header_idx]))
    if "name" not in cmap:
        return

    for row_cells, row_values in zip(rows[header_idx + 1:], plain_rows[header_idx + 1:]):
        if _is_divider_row(tuple(row_values)):
            continue
        name = row_values[cmap["name"]] if cmap["name"] < len(row_values) else None
        if not name or not str(name).strip():
            continue

        def get(key):
            i = cmap.get(key)
            if i is None or i >= len(row_values):
                return None
            v = row_values[i]
            return str(v).strip() if v not in (None, "") else None

        institution = get("company") or filename_institution
        if not institution:
            continue  # can't place this contact anywhere — skip rather than guess

        linkedin_url = None
        for key in ("linkedin", "name"):
            i = cmap.get(key)
            if i is not None and i < len(row_cells):
                hl = row_cells[i].hyperlink
                if hl and hl.target:
                    linkedin_url = hl.target
                    break

        tier_label = get("tier")
        notes_parts = [p for p in [
            f"[LinkedIn Priority: {tier_label}]" if tier_label else None,
            get("notes"),
            f"Location/company (raw): {get('location')}" if get("location") else None,
        ] if p]

        yield {
            "full_name": str(name).strip(),
            "institution": institution,
            "title": get("title"),
            "email": get("email"),
            "phone": get("phone"),
            "whatsapp": get("whatsapp"),
            "linkedin_url": linkedin_url,
            "seniority": get("seniority"),
            "persona": get("persona"),
            "country": get("country"),
            "notes": " | ".join(notes_parts) if notes_parts else None,
            "priority_score": TIER_PRIORITY_SCORE.get((tier_label or "").strip().lower()),
            "priority_tier": tier_label,
            "is_indian_origin": is_indian_sheet,
            "source": source_label,
        }


def _iter_csv_contacts(path: Path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = [list(r) for r in csv.reader(f)]
    if not rows:
        return
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return
    cmap = _col_map(tuple(rows[header_idx]))
    if "name" not in cmap:
        return
    institution_fallback = _institution_for_file(path)
    for row_values in rows[header_idx + 1:]:
        if _is_divider_row(tuple(row_values)):
            continue
        name = row_values[cmap["name"]] if cmap["name"] < len(row_values) else None
        if not name or not str(name).strip():
            continue

        def get(key):
            i = cmap.get(key)
            if i is None or i >= len(row_values):
                return None
            v = row_values[i]
            return str(v).strip() if v not in (None, "") else None

        institution = get("company") or institution_fallback
        if not institution:
            continue
        tier_label = get("tier")
        yield {
            "full_name": str(name).strip(), "institution": institution, "title": get("title"),
            "email": get("email"), "phone": get("phone"), "whatsapp": get("whatsapp"),
            "linkedin_url": get("linkedin"), "seniority": get("seniority"), "persona": get("persona"),
            "country": get("country"),
            "notes": f"[LinkedIn Priority: {tier_label}] {get('notes') or ''}".strip() if tier_label else get("notes"),
            "priority_score": TIER_PRIORITY_SCORE.get((tier_label or "").strip().lower()),
            "priority_tier": tier_label,
            "is_indian_origin": "india" in path.stem.lower(),
            "source": path.name,
        }


class _ContactUpserter:
    """Shared create-or-update core for a batch of contact dicts (as yielded by
    _iter_sheet_contacts / _iter_csv_contacts). Used by both the CLI (looping over
    every file in incoming/contacts/) and the dashboard's one-click 'process this
    upload now' action (a single in-memory file) — factored out so both paths run
    the exact same matching/upsert logic and can't silently drift apart."""

    def __init__(self, db):
        self.db = db
        self.created_orgs = 0
        self.created_people = 0
        self.updated_people = 0
        self.skipped = 0
        self.org_cache: dict[str, str] = {}
        # (org_id, full_name.lower()) -> Person, populated lazily. Needed because the
        # session has autoflush=False: without this cache, a person appearing in both
        # the main sheet AND an "Indians at X" / "India-located" subset sheet of the
        # SAME file would be inserted twice, since db.query() can't see uncommitted
        # rows added earlier in the same transaction.
        self.person_cache: dict[tuple[str, str], object] = {}

    def get_or_create_org(self, name: str) -> str:
        key = name.strip().lower()
        if key in self.org_cache:
            return self.org_cache[key]
        org = self.db.query(models.Organization).filter(models.Organization.canonical_name.ilike(name)).first()
        if not org:
            org = models.Organization(canonical_name=name, source="incoming/contacts import",
                                       verification_status="unverified")
            self.db.add(org)
            self.db.flush()
            self.db.add(models.OrgTypeTag(org_id=org.id, type_tag="commercial_bank"))
            self.created_orgs += 1
        self.org_cache[key] = org.id
        return org.id

    def get_existing_person(self, org_id: str, full_name: str):
        key = (org_id, full_name.strip().lower())
        if key in self.person_cache:
            return self.person_cache[key]
        p = self.db.query(models.Person).filter(
            models.Person.full_name.ilike(full_name), models.Person.current_org_id == org_id,
        ).first()
        if p:
            self.person_cache[key] = p
        return p

    def upsert_all(self, contact_iter):
        for c in contact_iter:
            if not c.get("full_name") or not c.get("institution"):
                self.skipped += 1
                continue
            org_id = self.get_or_create_org(c["institution"])
            existing = self.get_existing_person(org_id, c["full_name"])
            seniority = c.get("seniority") or infer_seniority(c.get("title"))
            flags = infer_flags(c.get("title"), seniority)
            fields = dict(
                current_title=c.get("title"), primary_email=c.get("email"), phone=c.get("phone"),
                whatsapp=c.get("whatsapp"), linkedin_url=c.get("linkedin_url"),
                seniority_level=seniority, persona=c.get("persona"), country=c.get("country"),
                background_notes=c.get("notes"), priority_score=c.get("priority_score"),
                is_decision_maker=flags["is_decision_maker"], is_influencer=flags["is_influencer"],
                priority_tier=c.get("priority_tier"), is_indian_origin=c.get("is_indian_origin"),
            )
            fields = {k: v for k, v in fields.items() if v is not None}
            if existing:
                for k, v in fields.items():
                    setattr(existing, k, v)
                self.updated_people += 1
            else:
                person = models.Person(full_name=c["full_name"], current_org_id=org_id,
                                        data_source=c.get("source"), **fields)
                self.db.add(person)
                self.db.flush()  # assigns person.id, makes it visible to person_cache immediately
                self.person_cache[(org_id, c["full_name"].strip().lower())] = person
                self.created_people += 1
        self.db.commit()

    def summary(self) -> dict:
        return {"orgs_created": self.created_orgs, "people_created": self.created_people,
                "people_updated": self.updated_people,
                "rows_skipped_missing_required_fields": self.skipped}


def import_contacts(db) -> dict:
    files = list(CONTACTS_DIR.glob("*.xlsx")) + list(CONTACTS_DIR.glob("*.csv"))
    upserter = _ContactUpserter(db)

    for path in files:
        if path.suffix.lower() == ".csv":
            contact_iter = _iter_csv_contacts(path)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
            filename_institution = _institution_for_file(path)
            contact_iter = []
            for ws in wb.worksheets:
                contact_iter = list(contact_iter) + list(
                    _iter_sheet_contacts(ws, filename_institution, f"{path.name} / {ws.title}"))
        upserter.upsert_all(contact_iter)

    return {"files": [f.name for f in files], **upserter.summary()}


def import_contacts_from_bytes(db, file_bytes: bytes, filename: str, institution_hint: str | None) -> dict:
    """Same matching/parsing logic as import_contacts(), but against an in-memory
    file instead of something sitting in incoming/contacts/ — this is what powers
    the dashboard's 'Process now' button on an upload. institution_hint is the bank
    the uploader picked/typed on the dashboard, used when a sheet has no Company
    column of its own (the common case for LinkedIn exports)."""
    import io
    suffix = Path(filename).suffix.lower()
    upserter = _ContactUpserter(db)

    if suffix == ".csv":
        text = file_bytes.decode("utf-8-sig", errors="replace")
        rows = [list(r) for r in csv.reader(io.StringIO(text))]
        if not rows:
            return {"error": "Empty file.", **upserter.summary()}
        header_idx = _find_header_row(rows)
        if header_idx is None or "name" not in _col_map(tuple(rows[header_idx])):
            return {"error": "Couldn't find a 'Name' column — is this a contact list?", **upserter.summary()}
        cmap = _col_map(tuple(rows[header_idx]))

        def contact_iter():
            for row_values in rows[header_idx + 1:]:
                if _is_divider_row(tuple(row_values)):
                    continue
                name = row_values[cmap["name"]] if cmap["name"] < len(row_values) else None
                if not name or not str(name).strip():
                    continue

                def get(key):
                    i = cmap.get(key)
                    if i is None or i >= len(row_values):
                        return None
                    v = row_values[i]
                    return str(v).strip() if v not in (None, "") else None

                institution = get("company") or institution_hint
                if not institution:
                    continue
                tier_label = get("tier")
                yield {
                    "full_name": str(name).strip(), "institution": institution, "title": get("title"),
                    "email": get("email"), "phone": get("phone"), "whatsapp": get("whatsapp"),
                    "linkedin_url": get("linkedin"), "seniority": get("seniority"), "persona": get("persona"),
                    "country": get("country"),
                    "notes": f"[LinkedIn Priority: {tier_label}] {get('notes') or ''}".strip() if tier_label else get("notes"),
                    "priority_score": TIER_PRIORITY_SCORE.get((tier_label or "").strip().lower()),
                    "priority_tier": tier_label,
                    "is_indian_origin": "india" in filename.lower(),
                    "source": filename,
                }
        upserter.upsert_all(contact_iter())

    elif suffix in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
        any_sheet_read = False
        for ws in wb.worksheets:
            rows_before = upserter.created_people + upserter.updated_people
            contact_list = list(_iter_sheet_contacts(ws, institution_hint, f"{filename} / {ws.title}"))
            if contact_list:
                any_sheet_read = True
            upserter.upsert_all(contact_list)
        if not any_sheet_read and upserter.skipped == 0:
            return {"error": "Couldn't find a contacts table in this file (no 'Name' column found on any sheet).",
                    **upserter.summary()}
    else:
        return {"error": f"Unsupported file type '{suffix}' — only .xlsx/.xlsm/.csv are auto-processed.",
                **upserter.summary()}

    return {"filename": filename, **upserter.summary()}


# ─────────────────────────────────────────────────────────────
#  VENDOR / SUBSIDIARY ECOSYSTEM IMPORT — one flat row per vendor/subsidiary,
#  auto-creates the org (if new) and its OrgRelationship link to the bank.
#  Shares get_or_create_org / upsert_org_relationship with the dashboard's
#  manual "add vendor" form so both paths dedupe identically — re-uploading
#  the same file (or hand-adding a vendor already in it) updates the existing
#  relationship instead of creating a second one.
# ─────────────────────────────────────────────────────────────
VENDOR_NAME_ALIASES = {"vendor_name", "vendor", "subsidiary_name", "subsidiary", "name", "organization", "company"}
VENDOR_TYPE_ALIASES = {"type", "vendor_type", "category", "classification"}
RELATIONSHIP_TYPE_ALIASES = {"relationship_type", "relationship", "connection_type", "role"}
STRENGTH_ALIASES = {"strength"}
CONFIDENCE_ALIASES = {"confidence", "confidence_%", "confidence_percent"}
CONTEXT_ALIASES = {"context", "notes", "description", "details"}
BANK_NAME_ALIASES = {"bank", "target_bank", "serves_bank", "bank_name"}


def get_or_create_org(db, name: str, default_type_tag: str = "vendor"):
    """Find an Organization by name (case-insensitive) or create it, tagged with
    default_type_tag if new. Returns (org, was_created). Shared by the ecosystem
    importer below and the dashboard's manual add-vendor route so a vendor typed
    by hand and one that arrives via file upload get deduped the same way."""
    org = db.query(models.Organization).filter(models.Organization.canonical_name.ilike(name)).first()
    if org:
        return org, False
    org = models.Organization(canonical_name=name, source="dashboard", verification_status="unverified")
    db.add(org)
    db.flush()
    db.add(models.OrgTypeTag(org_id=org.id, type_tag=default_type_tag))
    return org, True


def upsert_org_relationship(db, from_org_id: str, to_org_id: str, relationship_type: str = "vendor",
                             strength: str = "Weak", confidence: float | None = 0.5,
                             context: str | None = None, source: str | None = None):
    """Create-or-update the single OrgRelationship between this vendor/subsidiary and this
    bank. At most one relationship row is kept per (from_org_id, to_org_id) pair — re-running
    an import or re-adding the same vendor updates strength/confidence/context in place rather
    than creating a duplicate row in the ecosystem table. Returns (relationship, was_created)."""
    rel = db.query(models.OrgRelationship).filter(
        models.OrgRelationship.from_org_id == from_org_id,
        models.OrgRelationship.to_org_id == to_org_id,
    ).first()
    created = False
    if not rel:
        rel = models.OrgRelationship(from_org_id=from_org_id, to_org_id=to_org_id)
        db.add(rel)
        created = True
    rel.relationship_type = relationship_type or rel.relationship_type or "vendor"
    rel.strength = strength or rel.strength or "Weak"
    if confidence is not None:
        rel.confidence = confidence
    if context:
        rel.context = context
    if source:
        rel.source = source
    return rel, created


def _parse_confidence(raw: str | None) -> float:
    """Accepts '80', '80%', or '0.8' and normalizes to a 0-1 float. Falls back to 0.5
    (the model's own default) on anything unparseable rather than guessing higher/lower."""
    if not raw:
        return 0.5
    try:
        val = float(str(raw).strip().rstrip("%"))
        return val / 100 if val > 1 else val
    except ValueError:
        return 0.5


def import_ecosystem_from_bytes(db, file_bytes: bytes, filename: str, bank_hint: str | None) -> dict:
    """Reads a flat vendor/subsidiary list — one row per vendor, columns like Vendor Name /
    Type / Relationship Type / Strength / Confidence / Context (case-insensitive, only a name
    column is required) — and creates or updates each vendor's Organization plus its
    OrgRelationship link to the bank. bank_hint is the bank selected/typed on the dashboard
    upload, used when the file has no Bank/Company column of its own (the common case: one
    ecosystem file per bank, same shape as the per-bank contact-list uploads)."""
    import io
    suffix = Path(filename).suffix.lower()

    if suffix == ".csv":
        rows = list(csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig", errors="replace"))))
    elif suffix in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            return {"error": "Empty file."}
        headers = [str(h).strip() if h else "" for h in raw_rows[0]]
        rows = [dict(zip(headers, r)) for r in raw_rows[1:] if any(r)]
    else:
        return {"error": f"Unsupported file type '{suffix}' — only .xlsx/.xlsm/.csv are auto-processed."}

    if not rows:
        return {"error": "No data rows found in this file."}

    def _get(lower_row, alias_set):
        for alias in alias_set:
            if alias in lower_row and lower_row[alias] not in (None, ""):
                return str(lower_row[alias]).strip()
        return None

    vendors_created = 0
    relationships_updated = 0
    banks_created = 0
    skipped = 0
    bank_cache: dict[str, str] = {}
    any_name_column = False

    for raw in rows:
        lower_row = {str(k).strip().lower().replace(" ", "_"): v for k, v in raw.items() if k}
        vendor_name = _get(lower_row, VENDOR_NAME_ALIASES)
        if vendor_name:
            any_name_column = True
        if not vendor_name:
            skipped += 1
            continue
        bank_name = _get(lower_row, BANK_NAME_ALIASES) or bank_hint
        if not bank_name:
            skipped += 1
            continue

        bank_key = bank_name.strip().lower()
        if bank_key in bank_cache:
            bank_org_id = bank_cache[bank_key]
        else:
            bank_org, bank_was_created = get_or_create_org(db, bank_name, default_type_tag="commercial_bank")
            if bank_was_created:
                banks_created += 1
            bank_org_id = bank_org.id
            bank_cache[bank_key] = bank_org_id

        vendor_type_raw = _get(lower_row, VENDOR_TYPE_ALIASES)
        vendor_type_tag = vendor_type_raw.strip().lower().replace(" ", "_") if vendor_type_raw else "vendor"
        vendor_org, vendor_was_created = get_or_create_org(db, vendor_name, default_type_tag=vendor_type_tag)
        if vendor_was_created:
            vendors_created += 1
        elif vendor_type_raw:
            existing_tags = {t.type_tag for t in vendor_org.type_tags}
            if vendor_type_tag not in existing_tags:
                db.add(models.OrgTypeTag(org_id=vendor_org.id, type_tag=vendor_type_tag))

        relationship_type = (_get(lower_row, RELATIONSHIP_TYPE_ALIASES) or "vendor").strip().lower().replace(" ", "_")
        strength = _get(lower_row, STRENGTH_ALIASES) or "Weak"
        confidence = _parse_confidence(_get(lower_row, CONFIDENCE_ALIASES))
        context = _get(lower_row, CONTEXT_ALIASES)

        _, rel_created = upsert_org_relationship(
            db, from_org_id=vendor_org.id, to_org_id=bank_org_id,
            relationship_type=relationship_type, strength=strength,
            confidence=confidence, context=context, source=filename,
        )
        if not rel_created:
            relationships_updated += 1

    if not any_name_column and skipped == len(rows):
        db.rollback()
        return {"error": "Couldn't find a vendor/subsidiary name column in this file."}

    db.commit()
    return {
        "filename": filename, "vendors_created": vendors_created,
        "relationships_updated": relationships_updated,
        "banks_created": banks_created, "rows_skipped_missing_required_fields": skipped,
    }


OFFERING_HEADER_ALIASES = {
    "name": {"name", "product", "product_name", "offering"},
    "category": {"category", "type"},
    "description": {"description", "summary"},
    "target_segments": {"target_segments", "segments", "target_segment"},
    "key_benefits": {"key_benefits", "benefits"},
    "competitors": {"competitors", "competition"},
}


def _read_flat_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    elif path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h else "" for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(row)]
    return []


def _normalize_flat(row: dict, aliases: dict) -> dict:
    lower_row = {str(k).strip().lower().replace(" ", "_"): v for k, v in row.items() if k}
    out = {}
    for field, alias_set in aliases.items():
        for alias in alias_set:
            if alias in lower_row and lower_row[alias] not in (None, ""):
                out[field] = str(lower_row[alias]).strip()
                break
    return out


def _import_pdf_offering(db, path: Path) -> str:
    """One product per PDF: name = first non-empty line of text, description =
    everything else (trimmed). Good enough for a product-overview one-pager;
    category/target_segments/competitors are left blank for manual review."""
    try:
        import pdfplumber
    except ImportError:
        return "skipped_pdfplumber_not_installed"
    with pdfplumber.open(path) as pdf:
        text_pages = [p.extract_text() or "" for p in pdf.pages]
    full_text = "\n".join(text_pages).strip()
    lines = [l.strip() for l in full_text.splitlines() if l.strip()]
    if not lines:
        return "skipped_no_text"
    name = lines[0]
    description = "\n".join(lines[1:])[:8000]  # cap length; full text still in the PDF itself

    existing = db.query(models.Product).filter(models.Product.name.ilike(name)).first()
    if existing:
        existing.description = description or existing.description
        db.commit()
        return "updated"
    db.add(models.Product(name=name, description=description, category=None))
    db.commit()
    return "created"


def import_offerings(db) -> dict:
    files = list(OFFERINGS_DIR.glob("*.xlsx")) + list(OFFERINGS_DIR.glob("*.csv"))
    pdf_files = list(OFFERINGS_DIR.rglob("*.pdf"))
    pdf_created = pdf_updated = pdf_skipped = 0
    pdf_dep_missing = False
    for pdf_path in pdf_files:
        result = _import_pdf_offering(db, pdf_path)
        if result == "created":
            pdf_created += 1
        elif result == "updated":
            pdf_updated += 1
        elif result == "skipped_pdfplumber_not_installed":
            pdf_skipped += 1
            pdf_dep_missing = True
        else:
            pdf_skipped += 1

    created = updated = skipped = 0
    for path in files:
        for raw in _read_flat_rows(path):
            r = _normalize_flat(raw, OFFERING_HEADER_ALIASES)
            if not r.get("name"):
                skipped += 1
                continue
            existing = db.query(models.Product).filter(models.Product.name.ilike(r["name"])).first()
            target_segments = [s.strip() for s in r.get("target_segments", "").split(",") if s.strip()]
            competitors = [s.strip() for s in r.get("competitors", "").split(",") if s.strip()]
            if existing:
                existing.category = r.get("category") or existing.category
                existing.description = r.get("description") or existing.description
                existing.key_benefits = r.get("key_benefits") or existing.key_benefits
                if target_segments:
                    existing.target_segments = target_segments
                if competitors:
                    existing.competitors = competitors
                updated += 1
            else:
                db.add(models.Product(name=r["name"], category=r.get("category"),
                                       description=r.get("description"), key_benefits=r.get("key_benefits"),
                                       target_segments=target_segments, competitors=competitors))
                created += 1
        db.commit()

    non_tabular = [p for p in OFFERINGS_DIR.rglob("*")
                   if p.is_file() and p.suffix.lower() not in (".xlsx", ".csv", ".md", ".pdf")]
    return {"files": [f.name for f in files], "products_created": created,
            "products_updated": updated, "rows_skipped_missing_name": skipped,
            "pdf_files_found": [str(p.relative_to(OFFERINGS_DIR)) for p in pdf_files],
            "pdf_products_created": pdf_created, "pdf_products_updated": pdf_updated,
            "pdf_skipped_no_text": pdf_skipped,
            "pdf_dependency_missing": pdf_dep_missing,
            "other_files_found_not_imported": [str(p.relative_to(OFFERINGS_DIR)) for p in non_tabular]}


def main():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        print("=== Importing contacts from incoming/contacts/ ===")
        c = import_contacts(db)
        for k, v in c.items():
            print(f"  {k}: {v}")

        print("\n=== Importing offerings from incoming/offerings/ ===")
        o = import_offerings(db)
        for k, v in o.items():
            print(f"  {k}: {v}")
        if o.get("pdf_dependency_missing"):
            print("\n  NOTE: pdfplumber isn't installed, so PDF offerings were skipped.")
            print("  Run: pip install -r requirements.txt   then re-run this script.")
        if o["other_files_found_not_imported"]:
            print("\n  NOTE: found file(s) in incoming/offerings/ that were NOT imported")
            print("  (only .xlsx/.csv/.pdf are read):")
            for f in o["other_files_found_not_imported"]:
                print(f"    - {f}")
            print("  Tell Claude about these — likely a .zip that needs unpacking first.")

        db.add(models.AuditLog(action="import_incoming", details=str({**c, **o}), actor="import_incoming.py"))
        db.commit()
    finally:
        db.close()


if __name__ == "__main__":
    main()
